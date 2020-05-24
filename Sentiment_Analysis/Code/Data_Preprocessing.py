# @Date    : 16:52 05/01/2020
# @Author  : ClassicalPi
# @FileName: Data_Preprocessing.py
# @Software: PyCharm

import numpy
import pandas
import openpyxl
import os


# -*- coding: UTF-8 -*-

def getStaticInformation(City: str):
    os.chdir("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Data/{}".format(City))
    files = os.listdir("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Data/{}".format(City))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "{}".format(City)
    ws.append(["餐厅名称", "评论数量"])
    for file in files:
        if file.endswith(".xlsx"):
            print(file)
            wb2 = openpyxl.load_workbook(file)
            ws2 = wb2.active
            file = file.split('.')[0]
            name = file.replace("_", " ")
            number = ws2.max_row - 1
            ws.append([name, number])
            wb2.close()
    wb.save("{}.xlsx".format("{}_Static_Information".format(City)))
    print("Done")


def dateTransfer(old: str):
    # October 31, 2016
    try:
        month_day, year = old.split(", ")
        month, day = month_day.split(" ")
        return "{}-{}-{}".format(day, month, year)
    except:
        if old == "3 weeks ago":
            return "{}-{}-{}".format(22, "March", 2020)
        if old == "4 weeks ago":
            return "{}-{}-{}".format(15, "March", 2020)
        if old == "1 week ago":
            return "{}-{}-{}".format(7, "April", 2020)
        if old == "2 weeks ago":
            return "{}-{}-{}".format(14, "April", 2020)
        else:
            print(old)
            month = input("Month:")
            day = input("Day:")
            return "{}-{}-{}".format(day, month, 2020)


def toCSV(city: str):
    ans = ""
    os.chdir("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Data/{}".format(city))
    wb = openpyxl.load_workbook("All_{}.xlsx".format(city))
    ws = wb.active

    file = open("All_{}.csv".format(city), 'w')
    file.write("{},{},{},{},{},{}\n".format("ID", "Username", "City", "Country", "Date", "Comment"))
    index = 0
    for row in range(2, ws.max_row + 1):
        index += 1
        ans = ""
        ans += "{},".format(index)
        for i in range(1, 6):
            if i != 5:
                if i == 4:
                    date = str(ws.cell(row=row, column=i).value)
                    ans += dateTransfer(date)
                    ans += ','
                else:
                    ans += str(ws.cell(row=row, column=i).value)
                    ans += ","
            else:
                temp = str(ws.cell(row=row, column=i).value)
                if ord(temp[0]) >= 65 and ord(temp[0]) <= 122:
                    # temp.replace('\n',' ')
                    temp = temp.replace("\"", " ")
                    temp = "\"{}\"\n".format(temp)
                    ans += temp
                    file.write(ans)
                else:
                    index -= 1
    file.close()


def toExcel_ALL(citys: []):
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "ALL"
    ws1.append(["ID", "Username", "City", "Country", "review_date", "review_body"])
    index = 0
    for city in citys:
        os.chdir("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Data/{}".format(city))
        wb = openpyxl.load_workbook("All_{}.xlsx".format(city))
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            index += 1
            ans = []
            ans.append(index)
            for i in range(1, 6):
                if i != 5:
                    if i == 4:
                        date = str(ws.cell(row=row, column=i).value)
                        ans.append(date)
                    else:
                        ans.append(str(ws.cell(row=row, column=i).value))
                else:
                    temp = str(ws.cell(row=row, column=i).value)
                    if ord(temp[0]) >= 65 and ord(temp[0]) <= 122:
                        # temp.replace('\n',' ')
                        ans.append(temp)
                        ws1.append(ans)
                    else:
                        index -= 1
    wb1.save("All.xlsx")


def getSemester(url: str):
    wb = openpyxl.load_workbook("{}".format(url))
    ws = wb.active
    dict = {}
    for row in range(2, ws.max_row + 1):
        date = str(ws.cell(row=row, column=5).value)
        ans = dateTransfer(date)
        day, month, year = ans.split('-')
        if dict.__contains__(month):
            dict[month] += 1
        else:
            dict.setdefault(month, 1)
    print(dict)
    return dict


def toCSV_ALL(citys: []):
    file = open("All.csv", 'w')
    file.write("{},{},{},{},{},{}\n".format("ID", "Username", "City", "Country", "review_date", "review_body"))
    index = 0
    for city in citys:
        os.chdir("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Data/{}".format(city))
        wb = openpyxl.load_workbook("All_{}.xlsx".format(city))
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            index += 1
            ans = ""
            ans += "{},".format(index)
            for i in range(1, 6):
                if i != 5:
                    if i == 4:
                        date = str(ws.cell(row=row, column=i).value)
                        ans += dateTransfer(date)
                        ans += ','
                    else:
                        ans += str(ws.cell(row=row, column=i).value)
                        ans += ","
                else:
                    temp = str(ws.cell(row=row, column=i).value)
                    if ord(temp[0]) >= 65 and ord(temp[0]) <= 122:
                        # temp.replace('\n',' ')
                        temp = temp.replace("\"", " ")
                        temp = "\"{}\"\n".format(temp)
                        ans += temp
                        file.write(ans)
                    else:
                        index -= 1
    file.close()


if __name__ == '__main__':
    # toCSV("Macau")
    # toExcel_ALL(["Hongkong","GuangZhou","Macau"])
    getSemester("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Code/All.xlsx")
