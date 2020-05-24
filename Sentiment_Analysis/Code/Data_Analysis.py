# @Date    : 22:38 05/01/2020
# @Author  : ClassicalPi
# @FileName: Data_Analysis.py
# @Software: PyCharm

import numpy as np

from pyecharts.charts import Map,Geo
from pyecharts import options
import pandas as pd
import nltk
import re
import os
import matplotlib.pyplot as plt
import string
import json
import openpyxl
from nltk.stem.snowball import SnowballStemmer



def loadReview(city: str, res: str or None, All: False) -> str:
    ans = ""
    os.chdir("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Data/{}".format(city))
    if All:
        wb = openpyxl.load_workbook("All_{}.xlsx".format(city))
    else:
        wb = openpyxl.load_workbook("{}".format(res))
    ws = wb.active
    for row in range(1, ws.max_row):
        temp = str(ws.cell(row=row, column=5).value)
        if ord(temp[0]) >= 65 and ord(temp[0]) <= 122:
            ans += temp
            ans += '\n'
    return ans


def load_excel_url(url: str):
    wb = openpyxl.load_workbook("{}".format(url))
    ws = wb.active
    ans = ""
    for row in range(2, ws.max_row):
        temp = str(ws.cell(row=row, column=6).value)
        if ord(temp[0]) >= 65 and ord(temp[0]) <= 122:
            ans += temp
            ans += '\n'
    return ans


def tokenize_and_stem(text):
    stemmer = SnowballStemmer("english")
    # first tokenize by sentence, then by word to ensure that punctuation is caught as it's own token
    tokens = [word for sent in nltk.sent_tokenize(text) for word in nltk.word_tokenize(sent)]
    filtered_tokens = []
    # filter out any tokens not containing letters (e.g., numeric tokens, raw punctuation)
    for token in tokens:
        if re.search('[a-zA-Z]', token):
            filtered_tokens.append(token)
    stems = [stemmer.stem(t) for t in filtered_tokens]
    return stems


def tokenize_only(text):
    # first tokenize by sentence, then by word to ensure that punctuation is caught as it's own token
    tokens = [word.lower() for sent in nltk.sent_tokenize(text) for word in nltk.word_tokenize(sent)]
    filtered_tokens = []
    # filter out any tokens not containing letters (e.g., numeric tokens, raw punctuation)
    for token in tokens:
        if re.search('[a-zA-Z]', token):
            filtered_tokens.append(token)
    return filtered_tokens


def drawBarh(dic: dict, num: int):
    """
    绘制水平条形图方法barh
    参数一：数据字典
    参数二：topN
    """
    listkey = []
    listval = []
    for key, val in sorted(dic.items(), key=lambda x: (x[1], x[0]), reverse=True)[:num]:
        listkey.append(key)
        listval.append(val)
    df = pd.DataFrame(listval[::-1], columns=[u'Times'])
    df.index = listkey[::-1]
    df.plot(kind='barh', color="lightblue")
    plt.title(u'Top {} Most Common Words in Reviews of Hongkong,GuangZhou and Macau'.format(num))
    plt.show()


def tf_idf(city: str):
    from sklearn.feature_extraction.text import TfidfVectorizer
    from nltk.stem.porter import PorterStemmer

    path = '../Data/{}/'.format(city)
    token_dict = {}

    def tokenize(text):
        tokens = nltk.word_tokenize(text)
        stems = []
        for item in tokens:
            stems.append(item)
        return stems

    for dirpath, dirs, files in os.walk(path):
        for f in files:
            fname = os.path.join(dirpath, f)
            print("fname=", fname)
            text = loadReview(city, All=False, res=f)
            token_dict[f] = text.lower().translate(string.punctuation)
    tfidf = TfidfVectorizer(tokenizer=tokenize, stop_words='english')
    tfs = tfidf.fit_transform(token_dict.values())

    str = loadReview("GuangZhou", All=True, res=None)
    response = tfidf.transform([str])
    print(response)

    feature_names = tfidf.get_feature_names()
    for col in response.nonzero()[1]:
        print(feature_names[col], ' - ', response[0, col])


def getFrequency(text: str, topN: int):
    stopwords = nltk.corpus.stopwords.words('english')

    stopwords.append('\'s')
    stopwords.append('n\'t')
    stopwords.append('us')
    stopwords.append("restaurant")
    stopwords.append("one")
    stopwords.append("ordered")
    stopwords.append("try")
    stopwords.append("sum")
    stopwords.append("kong")

    totalvocab_tokenized = []
    totalvocab_tokenized_stem = []
    for i in [text]:
        allwords_tokenized = tokenize_only(i)
        totalvocab_tokenized.extend(allwords_tokenized)
        totalvocab_tokenized_stem.extend(tokenize_and_stem(i))
    words = []
    transfer = {
        "dim": "dim sum",
        "hong": "hong kong"
    }
    for each in totalvocab_tokenized:
        if each not in stopwords:
            if each in transfer:
                words.append(transfer[each])
            else:
                words.append(each)
    fre = nltk.FreqDist(words)
    drawBarh(fre, topN)

def getinfo(word:str):
    wb = openpyxl.load_workbook("{}".format("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Code/All.xlsx"))
    ws = wb.active
    ans=0
    for row in range(2, ws.max_row + 1):
        comment = str(ws.cell(row=row, column=6).value)
        comment=comment.lower()
        if word in comment:
            ans+=1
            continue
    return ans
def category():
    base=open("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Code/AFINN.csv",'r')
    base.readline()
    dic={}
    for eachline in base.readlines():
        word,num=eachline.split(";")
        num=int(num.strip(' '))
        if dic.__contains__(word):
            continue
        else:
            dic.setdefault(word,num)

    file=open("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Code/category.csv",'r')
    ans={}
    file.readline()
    for eachline in file.readlines():
        decor,cate,num=eachline.split(';')
        num=num.strip(' ')
        if ans.__contains__(cate):
            if dic.__contains__(decor):
                if ans[cate].__contains__(dic[decor]):
                    ans[cate][dic[decor]]+=1
                    ans[cate]["Sum"]+=int(dic[decor])
                    if int(dic[decor])>=0:
                        ans[cate]["Positive Count"]+=1
                        ans[cate]["Positive Score"] += int(dic[decor])
                    else:
                        ans[cate]["Negative Count"]+=1
                        ans[cate]["Negative Score"] += int(dic[decor])
                else:
                    ans[cate].setdefault(dic[decor],1)
                    ans[cate]["Sum"] += int(dic[decor])
                    if int(dic[decor])>=0:
                        ans[cate]["Positive Count"]+=1
                        ans[cate]["Positive Score"] += int(dic[decor])
                    else:
                        ans[cate]["Negative Count"]+=1
                        ans[cate]["Negative Score"] += int(dic[decor])
            else:
                continue
        else:
            if dic.__contains__(decor):
                ans.setdefault(cate,{dic[decor]:1})
                ans[cate].setdefault("Sum",int(dic[decor]))
                if int(dic[decor])>0:
                    ans[cate].setdefault("Positive Count",1)
                    ans[cate].setdefault("Positive Score",int(dic[decor]))
                    ans[cate].setdefault("Negative Count",0)
                    ans[cate].setdefault("Negative Score",0)
                else:
                    ans[cate].setdefault("Negative Count",1)
                    ans[cate].setdefault("Negative Score",int(dic[decor]))
                    ans[cate].setdefault("Positive Count",0)
                    ans[cate].setdefault("Positive Score", 0)
            else:
                continue

    print(ans)
    file.close()
    with open('category2.json', 'w') as fp:
        json.dump(ans,fp)
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "ALL"
    trans={
        -5:"1",
        -4:"2",
        -3:"3",
        -2:"4",
        -1:"5",
        5: "10",
        4: "9",
        3: "8",
        2: "7",
        1: "6",
        "Negative Count":"11",
        "Negative Score": "12",
        "Positive Count": "13",
        "Positive Score": "14",
        "Sum":15
    }
    ws1.append(["Word", "痛恨：-5", "非常讨厌：-4", "很讨厌：-3", "比较讨厌：-2",
                "不适：-1","有好感：1","比较喜欢：2","喜欢：3","很喜欢：4","非常喜欢：5",
                "负面情感数量占比","负面情感分数占比","正面情感数量占比","正面情感分数分数","情感分数总计"])
    for word,w_v in ans.items():
        temp = [0 for a in range(0,16)]
        temp[0]=word
        for num,count in w_v.items():
            index=int(trans[num])
            temp[index]=count
        total_amount=temp[11]+temp[13]
        temp[11]=temp[11]/total_amount
        temp[13] = temp[13] / total_amount
        temp[15]=abs(temp[12])+temp[14]
        if temp[15]!=0:
            abs_score=abs(temp[12])+temp[14]
            temp[12]=abs(temp[12])/abs_score
            temp[14] = temp[14]/abs_score
        ws1.append(temp)
    wb1.save("category.xlsx")

def draw_heatmap():
    wb = openpyxl.load_workbook("{}".format("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Code/All.xlsx"))
    ws = wb.active
    states = {"Alabama", "Alaska", "Arizona", "Arkansas", "California"
        , "Colorado"
        , "Connecticut"
        , "Delaware"
        , "Florida"
        , "Georgia"
        , "Hawaii"
        , "Idaho"
        , "Illinois"
        , "Indiana"
        , "Iowa"
        , "Kansas"
        , "Kentucky"
        , "Louisiana"
        , "Maine"
        , "Maryland"
        , "Massachusetts"
        , "Michigan"
        , "Minnesota"
        , "Mississippi"
        , "Missouri"
        , "MontanaNebraska"
        , "Nevada"
        , "New Hampshire"
        , "New Jersey"
        , "New Mexico"
        , "New York"
        , "North Carolina"
        , "North Dakota"
        , "Ohio"
        , "Oklahoma"
        , "Oregon"
        , "PennsylvaniaRhode Island"
        , "South Carolina"
        , "South Dakota"
        , "Tennessee"
        , "Texas"
        , "Utah"
        , "Vermont"
        , "Virginia"
        , "Washington"
        , "West Virginia"
        , "Wisconsin"
        , "Wyoming"
              }
    dict = {}
    for row in range(2, ws.max_row + 1):
        country = str(ws.cell(row=row, column=4).value)
        if country != "None":
            try:
                if country in states:
                    country="United States"
                if dict.__contains__(country):
                    dict[country]+=1
                else:
                    dict.setdefault(country,1)
            except:
                print(country)
    c=[]
    u=[]
    for k,v in dict.items():
        c.append(k)
        u.append(int(v))
    list1 = [[c[i], u[i]] for i in range(len(c))]
    map_1 = Map().add(
        series_name="游客数量",  # 名称
        data_pair=list1,  # 传入数据
        is_map_symbol_show=False,
        maptype='world',  # 地图类型
    )
    map_1.set_series_opts(label_opts=options.LabelOpts(is_show=False))
    map_1.set_global_opts(
        title_opts=options.TitleOpts(title="粤港澳境外游客 客源地热力图"),
        visualmap_opts=options.VisualMapOpts(
            max_=1500,
            is_piecewise=True,
            pieces=[
            {"min": 1000,"color": '#000000'},
            {"min": 600, "max": 1000,"color": '#4f4f4f'},
            {"min": 300, "max": 600,"color": '#757575'},
            {"min": 100, "max": 300,"color": '#919191'},
            {"min": 10, "max": 100,"color": '#adadad'},
            {"max": 10,"color": '#d1d1d1'}, ]
        )
    )
    map_1.render('map4.html')


if __name__ == '__main__':
    #synopses = [load_excel_url("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Code/All.xlsx")]
    #getFrequency(synopses[0], 20)
    # tf_idf("HongKong")
    # draw_heatmap()
    #for w in ['view','views','atmosphere','decor','decoration','toilet','environment']:
    #    print("{}:{}".format(w,getinfo(w)))
    category()